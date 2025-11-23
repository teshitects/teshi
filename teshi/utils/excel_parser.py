from pathlib import Path
import openpyxl
import xlrd


class ExcelParser:
    def __init__(self, file_path):
        self.file_path = file_path
        self.rows = []

    def parse(self):
        suffix = Path(self.file_path).suffix.lower()
        if suffix == ".xlsx":
            self._parse_xlsx()
        elif suffix == ".xls":
            self._parse_xls()

    def _parse_xlsx(self):
        wb = openpyxl.load_workbook(self.file_path)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            self.rows.append(list(row))


    def _parse_xls(self):
        # openpyxl don't support xls, use xlrd
        wb = xlrd.open_workbook(self.file_path)
        ws = wb.sheet_by_index(0)
        for row in ws.iter_rows(values_only=True):
            self.rows.append(list(row))